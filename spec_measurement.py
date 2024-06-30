import ipaddress
import random
from tkinter import *
from tkinter import ttk
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import numpy as np
from zolix.app.zolix_gateway import ZolixGateway
from RigolLib import RigolLib
from openpyxl import Workbook
from openpyxl.chart import ScatterChart, Reference, Series
from tkinter import filedialog
from datetime import datetime


class SpectralMeasurements:
    """Class for creating window to measure spectre of the sample using oscilloscope and monochromator"""

    def __init__(self):
        self.zolix_gateway = None
        self.zolix_connected = False
        self.rigol_gateway = RigolLib.Scope()
        self.rigol_connected = False
        self.zolix_IP_value = "127.0.0.1"
        self.oscilloscope_chanel = "ch1"

        # Первоначальные данные графика
        self.x_values = []
        self.y_values = []

        self.root = Tk()  # create new tkinter obj
        self._create_interface()
        self.root.mainloop()

    def _change_monochromator_wavelength(self, new_wavelength):
        # if self.zolix_gateway:
        #     self.zolix_gateway.move_to_wave(new_wavelength)  # set new wavelength
        #     cur_wave = self.zolix_gateway.get_current_wave()  # get current wavelength
        #     if cur_wave == new_wavelength:  # check whether the wavelength is set
        #         return True
        # return False
        return True

    def _check_all_equipment_connected(self):
        if self.rigol_connected and self.zolix_connected:
            self.wavelength_from_input.config(state="normal")
            self.wavelength_to_input.config(state="normal")
            self.wavelength_measurement_step_input.config(state="normal")
            self.start_measurement_button.config(state="normal")
            self.channels_selection_box.config(state="normal")

    def _connect_to_Rigol_oscilloscope(self):
        # self.rigol_gateway.
        # rigol_gateway.auto()
        # rigol_gateway.run()
        # self.rigol_gateway = rigol_gateway
        self.rigol_connected = True
        self._check_all_equipment_connected()
        self.rigol_connect_state.config(text="Подключено", background="#50FA1C")

    def _connect_to_Zolix_monochromator(self):
        """Method to connect to the zolix monochromator. Don't forger to connect usb and turn on the zolix server"""
        if self._validate_IP_zolix():
            # zolix_gateway = ZolixGateway(
            #     f"{self.zolix_IP.get()}", 43665
            # )  # configure ip address and port of the client
            # zolix_gateway.connect_to_server()  # connect to the server
            # zolix_gateway.set_usb_mode(True)  # set the mode of communication in USB mode
            # qte = (
            #     zolix_gateway.search_zolix_usb_device()
            # )  # search all zolix connected with the server
            # serial = zolix_gateway.get_zolix_usb_serial(0)
            # zolix_gateway.set_usb_serials(serial)

            # zolix_gateway.get_is_open()  # verify if there is an already connected monochromator
            # zolix_gateway.open()  # open the communication with the zolix monochromator and the server
            # zolix_gateway.get_is_open()  # verify that the connection between the server and the monochromator is on
            # self.zolix_gateway = zolix_gateway
            self.zolix_connected = True
            self._check_all_equipment_connected()
            self.zolix_connect_state.config(
                text="Подключено", foreground="#000000", background="#50FA1C"
            )
            return True
        else:
            self.zolix_connect_state.config(
                text="Неверный формат IP адреса.",
                foreground="#B71C1C",
                background="#F71E1E",
            )

    def _create_interface(self):
        root = self.root

        root.title("Спектрометр")  # give window name
        validate_float_only = (root.register(self._validate_wl_input_float_only), "%P")

        self.opts = {"padx": 10, "pady": 10, "ipadx": 10, "ipady": 10, "sticky": "nswe"}
        opts = self.opts

        # IP for Zolix
        self.zolix_IP = Entry()
        self.zolix_IP.insert(0, self.zolix_IP_value)
        self.zolix_IP.grid(row=0, column=0, **opts)

        # Zolix connect
        self.zolix_connect = Button(
            text="Подключить Zolix", command=self._connect_to_Zolix_monochromator
        )
        self.zolix_connect.grid(row=0, column=1, **opts)

        # Zolix connect state
        self.zolix_connect_state = Label(text="Отключено", background="#F71E1E")
        self.zolix_connect_state.grid(row=0, column=2, **opts)

        # Rigol USB Options
        self.rigol_usb_chosen = ttk.Combobox(
            values=self.rigol_gateway.get_available_usb_devices(),
        )

        # self.rigol_usb_chosen.current(0)
        self.rigol_usb_chosen.grid(row=1, column=0, **opts)
        self.rigol_usb_chosen.bind("<<ComboboxSelected>>", self._set_device_for_Rigol)

        # Rigol USB connect
        self.rigol_connect = Button(
            text="Подключить Rigol", command=self._connect_to_Rigol_oscilloscope
        )
        self.rigol_connect.grid(row=1, column=1, **opts)

        # Rigol connect state
        self.rigol_connect_state = Label(text="Отключено", background="#F71E1E")
        self.rigol_connect_state.grid(row=1, column=2, **opts)

        # Oscilloscope channel selection label
        self.channel_selection_label = Label(text="Канал осциллографа")
        self.channel_selection_label.grid(row=2, column=0, **opts)

        # Oscilloscope channel selection
        self.channels_selection_box = ttk.Combobox(
            state=DISABLED,
            values=["ch1", "ch2"],
        )
        self.channels_selection_box.current(0)
        self.channels_selection_box.grid(row=2, column=1, **opts)
        self.channels_selection_box.bind(
            "<<ComboboxSelected>>", self._set_oscilloscope_chanel
        )

        # label for initial wavelength
        self.wavelength_from_label = Label(text="Начальная длина волны")
        self.wavelength_from_label.grid(row=4, column=0, **opts)

        # input for initial wavelength
        self.wavelength_from_input = Entry(
            state=DISABLED, validate="key", validatecommand=validate_float_only
        )
        self.wavelength_from_input.grid(row=5, column=0, **opts)

        # label for final wavelength
        self.wavelength_to_label = Label(text="Конечная длина волны")
        self.wavelength_to_label.grid(row=4, column=1, **opts)

        # input for final wavelength
        self.wavelength_to_input = Entry(
            state=DISABLED,
            validate="key",
            validatecommand=validate_float_only,
        )
        self.wavelength_to_input.grid(row=5, column=1, **opts)

        # label for step of wavelengths
        self.wavelength_measurement_step_label = Label(text="Шаг измерения")
        self.wavelength_measurement_step_label.grid(row=4, column=2, **opts)

        # input for step of wavelengths
        self.wavelength_measurement_step_input = Entry(
            state=DISABLED, validate="key", validatecommand=validate_float_only
        )
        self.wavelength_measurement_step_input.grid(row=5, column=2, **opts)

        # start measurement button
        self.start_measurement_button = Button(
            state=DISABLED,
            text="Начать измерение",
            command=self._start_measurement,
        )
        self.start_measurement_button.grid(row=6, column=0, columnspan=3, **opts)

    def _get_plot_data(self):
        wb = Workbook()
        sheet = wb.active
        sheet.title = "Результаты измерения"
        sheet.cell(1, 1).value = "Длина волны"
        sheet.cell(1, 2).value = "Амплитуда"

        # Add data to the first column of the file
        for row in range(2, len(self.x_values) + 2):
            sheet.cell(row, 1).value = self.x_values[row - 2]

        # Add data to the first column of the file
        for row in range(2, len(self.y_values) + 2):
            sheet.cell(row, 2).value = self.y_values[row - 2]

        chart = ScatterChart()

        x_values = Reference(
            sheet, min_col=1, min_row=1, max_row=len(self.x_values) + 1
        )

        y_values = Reference(
            sheet, min_col=2, min_row=1, max_row=len(self.x_values) + 1
        )

        series1 = Series(y_values, x_values, title_from_data=True)

        chart.title = "Спектр излучения"
        chart.y_axis.tital = "Амплитуда, у.е"
        chart.x_axis.tital = "Длина волны, нм"

        values = Reference(
            worksheet=sheet,
            min_row=1,
            max_row=len(self.x_values) + 1,
            min_col=1,
            max_col=2,
        )

        chart.append(series1)

        sheet.add_chart(chart, "D2")

        # Open a save file dialog
        file_path = filedialog.asksaveasfilename(
            title="Выберите путь сохранения файла",
            initialfile=f"Spectrum_range:{self.initial_wl}-{self.final_wl}_step:{self.step}_{datetime.today().strftime('%d.%m.%Y')}.xlsx",
            filetypes=[("Файлы Excel", "*.xlsx")],
            defaultextension=".xlsx",
        )

        # Сохраняем файл Excel
        wb.save(file_path)

    def _get_Rigol_oscillograph_average_V(self):
        # if self.rigol_gateway:
        #     if self.oscilloscope_chanel == "ch1":
        #         return self.rigol_gateway.ch1.meas_Vavg()
        #     elif self.oscilloscope_chanel == "ch2":
        #         return self.rigol_gateway.ch2.meas_Vavg()
        return random.randint(1, 10)

    def _get_Rigol_oscillograph_max_V(self):
        # if self.rigol_gateway:
        #     if self.oscilloscope_chanel == "ch1":
        #         return self.rigol_gateway.ch1.meas_Vmax()
        #     elif self.oscilloscope_chanel == "ch2":
        #         return self.rigol_gateway.ch2.meas_Vmax()
        return 10

    def _get_Rigol_oscillograph_min_V(self):
        # if self.rigol_gateway:
        #     if self.oscilloscope_chanel == "ch1":
        #         return self.rigol_gateway.ch1.meas_Vmin()
        #     elif self.oscilloscope_chanel == "ch2":
        #         return self.rigol_gateway.ch2.meas_Vmin()
        return 0

    def _plot(self):
        # Очищаем прошлые данные
        self.x_values = []
        self.y_values = []

        # Включаем интерактивный режим
        plt.ion()

        # Создаем объект графика
        fig = plt.Figure()
        ax = fig.add_subplot(111)

        # Устанавливаем границы графика
        ax.set_xlim(self.initial_wl, self.final_wl)
        ax.set_ylim(
            self._get_Rigol_oscillograph_min_V(), self._get_Rigol_oscillograph_max_V()
        )

        # формируем список точек для измерения
        x_range = np.arange(self.initial_wl, self.final_wl + self.step, self.step)

        # Формируем первичную линию графика
        (line1,) = ax.plot(self.x_values, self.y_values, "b-")

        # Добавляем наш график в окно
        canvas = FigureCanvasTkAgg(fig, master=self.root)

        # Располагаем график в Tkinter окне
        canvas.get_tk_widget().grid(row=7, column=0, columnspan=3, **self.opts)

        # Пробегаемся по точкам измерения и получаем данные с приборов, и обновляем график
        for x in x_range:
            # if change_monochromator_wavelength(x):
            if self._change_monochromator_wavelength(x):
                self.x_values.append(float(x))
                self.y_values.append(self._get_Rigol_oscillograph_average_V())

                line1.set_xdata(self.x_values)
                line1.set_ydata(self.y_values)

                # Обновляем график
                fig.canvas.draw()
                fig.canvas.flush_events()

        # start measurement button
        self.get_plot_data_button = Button(
            self.root,
            text="Сохранить Excel файл",
            command=self._get_plot_data,
        )
        self.get_plot_data_button.grid(row=8, column=0, **self.opts)
        # Обновляем график
        fig.canvas.draw()
        fig.canvas.flush_events()

    def _set_device_for_Rigol(self, event):
        val = self.rigol_usb_chosen.get()
        self.rigol_device = val

    def _set_oscilloscope_chanel(self, event):
        val = self.channels_selection_box.get()
        self.oscilloscope_chanel = val

    def _validate_IP_zolix(self):
        try:
            ipaddress.IPv4Network(self.zolix_IP.get())
            return True
        except ValueError:
            return False

    def _validate_wl_input_float_only(self, string):
        if (
            string.isdigit()
            or (string and string.replace(".", "", 1).isdigit())
            or not string
        ):
            return True
        else:
            return False

    def _start_measurement(self):
        self.initial_wl = float(self.wavelength_from_input.get())
        self.final_wl = float(self.wavelength_to_input.get())
        self.step = float(self.wavelength_measurement_step_input.get())
        self.animationObj = self._plot()
